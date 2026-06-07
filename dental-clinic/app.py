# -*- coding: utf-8 -*-
"""
口腔诊所轻量化管理系统 - 主程序
技术栈: Flask + SQLite + Jinja2模板
适合部署在2G内存的轻量云服务器
"""
import os
import json
import sqlite3
from datetime import datetime, date
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, send_from_directory, flash, Response)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

app = Flask(__name__)
app.secret_key = 'dental-clinic-secret-key-2024'

# 自定义 Jinja2 过滤器
@app.template_filter('from_json')
def from_json_filter(s):
    try:
        return json.loads(s or '[]')
    except Exception:
        return []
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'bmp', 'webp'}

DB_PATH = os.path.join(os.path.dirname(__file__), 'dental.db')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# ============================================================
# 数据库初始化
# ============================================================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    """初始化数据库表结构"""
    conn = get_db()
    c = conn.cursor()

    # 用户表
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        real_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'doctor',
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )''')

    # 诊所基础信息
    c.execute('''CREATE TABLE IF NOT EXISTS clinic_settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')

    # 数据字典
    c.execute('''CREATE TABLE IF NOT EXISTS dict_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT NOT NULL,
        name TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0,
        is_active INTEGER DEFAULT 1
    )''')

    # 客户档案
    c.execute('''CREATE TABLE IF NOT EXISTS customers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        gender TEXT,
        birth_year INTEGER,
        phone TEXT UNIQUE NOT NULL,
        wechat TEXT,
        address TEXT,
        dental_history TEXT,
        allergy_history TEXT,
        systemic_disease TEXT,
        dental_condition TEXT,
        tags TEXT DEFAULT '[]',
        first_doctor_id INTEGER,
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (first_doctor_id) REFERENCES users(id)
    )''')

    # 诊疗记录
    c.execute('''CREATE TABLE IF NOT EXISTS treatments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id INTEGER NOT NULL,
        visit_time TEXT NOT NULL,
        doctor_id INTEGER NOT NULL,
        visit_type TEXT DEFAULT '初诊',
        chief_complaint TEXT,
        exam_result TEXT,
        xray_result TEXT,
        diagnosis TEXT,
        treatment_items TEXT DEFAULT '[]',
        treatment_area TEXT,
        treatment_process TEXT,
        drug_name TEXT,
        drug_dosage TEXT,
        drug_usage TEXT,
        drug_period TEXT,
        total_fee REAL DEFAULT 0,
        paid_fee REAL DEFAULT 0,
        unpaid_fee REAL DEFAULT 0,
        advice TEXT,
        contraindications TEXT,
        next_visit_date TEXT,
        next_exam_items TEXT,
        status TEXT DEFAULT 'draft',
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (customer_id) REFERENCES customers(id),
        FOREIGN KEY (doctor_id) REFERENCES users(id)
    )''')

    # 诊疗附件
    c.execute('''CREATE TABLE IF NOT EXISTS treatment_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        treatment_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        original_name TEXT NOT NULL,
        file_type TEXT,
        uploaded_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (treatment_id) REFERENCES treatments(id)
    )''')

    # 回访计划
    c.execute('''CREATE TABLE IF NOT EXISTS followups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id INTEGER NOT NULL,
        treatment_id INTEGER,
        plan_date TEXT NOT NULL,
        followup_type TEXT DEFAULT '术后回访',
        plan_content TEXT,
        status TEXT DEFAULT 'pending',
        actual_date TEXT,
        result TEXT,
        recovery TEXT,
        next_appointment INTEGER DEFAULT 0,
        next_followup_needs TEXT,
        created_by INTEGER,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (customer_id) REFERENCES customers(id),
        FOREIGN KEY (treatment_id) REFERENCES treatments(id)
    )''')

    # 操作日志
    c.execute('''CREATE TABLE IF NOT EXISTS operation_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        action TEXT,
        target_type TEXT,
        target_id INTEGER,
        detail TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )''')

    conn.commit()

    # 初始化默认管理员
    admin = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()
    if not admin:
        c.execute("INSERT INTO users (username, password_hash, real_name, role) VALUES (?,?,?,?)",
                  ('admin', generate_password_hash('admin123'), '管理员', 'admin'))

    # 初始化诊所设置
    settings = [
        ('clinic_name', '口腔诊所'),
        ('clinic_phone', ''),
        ('clinic_address', ''),
        ('clinic_slogan', '专业口腔 健康生活'),
    ]
    for key, val in settings:
        c.execute("INSERT OR IGNORE INTO clinic_settings(key,value) VALUES(?,?)", (key, val))

    # 初始化数据字典
    default_dicts = [
        ('treatment_item', ['补牙', '拔牙', '洁牙', '正畸', '种植牙', '根管治疗', '贴面', '牙冠', '牙周治疗', '口腔检查']),
        ('disease', ['龋齿', '牙周炎', '牙髓炎', '牙龈炎', '智齿冠周炎', '口腔溃疡', '牙列不齐', '牙本质过敏', '根尖周炎']),
        ('drug', ['阿莫西林', '甲硝唑', '布洛芬', '地塞米松', '氯己定漱口水', '碘仿糊剂', '氧化锌丁香油糊剂']),
        ('followup_type', ['术后回访', '复诊提醒', '常规问候', '正畸复查', '种植复查']),
        ('customer_tag', ['新客户', '老客户', '高需求客户', '正畸客户', '补牙客户', '种植客户', 'VIP客户', '待跟进']),
        ('visit_type', ['初诊', '复诊', '急诊', '复查', '咨询']),
    ]
    for cat, items in default_dicts:
        existing = conn.execute("SELECT COUNT(*) FROM dict_items WHERE category=?", (cat,)).fetchone()[0]
        if existing == 0:
            for i, name in enumerate(items):
                c.execute("INSERT INTO dict_items(category,name,sort_order) VALUES(?,?,?)", (cat, name, i))

    conn.commit()
    conn.close()


# ============================================================
# 权限辅助
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return jsonify({'success': False, 'msg': '权限不足，仅管理员可操作'}), 403
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    if 'user_id' in session:
        conn = get_db()
        u = conn.execute("SELECT * FROM users WHERE id=?", (session['user_id'],)).fetchone()
        conn.close()
        return u
    return None


def get_clinic_settings():
    conn = get_db()
    rows = conn.execute("SELECT key,value FROM clinic_settings").fetchall()
    conn.close()
    return {r['key']: r['value'] for r in rows}


def log_action(action, target_type=None, target_id=None, detail=None):
    try:
        conn = get_db()
        conn.execute("INSERT INTO operation_logs(user_id,action,target_type,target_id,detail) VALUES(?,?,?,?,?)",
                     (session.get('user_id'), action, target_type, target_id, detail))
        conn.commit()
        conn.close()
    except Exception:
        pass


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# 页面路由
# ============================================================
@app.route('/')
@login_required
def index():
    conn = get_db()
    today = date.today().isoformat()
    # 今日待回访
    today_followups = conn.execute(
        """SELECT f.*, c.name as customer_name, c.phone FROM followups f
           LEFT JOIN customers c ON f.customer_id=c.id
           WHERE f.plan_date=? AND f.status='pending' ORDER BY f.id""", (today,)
    ).fetchall()
    # 近3天待复诊
    upcoming_next_visits = conn.execute(
        """SELECT t.*, c.name as customer_name, c.phone FROM treatments t
           LEFT JOIN customers c ON t.customer_id=c.id
           WHERE t.next_visit_date BETWEEN ? AND date(?,'+3 days')
           AND t.status='done' ORDER BY t.next_visit_date""", (today, today)
    ).fetchall()
    # 本月统计
    month_start = today[:7] + '-01'
    stats = {
        'total_customers': conn.execute("SELECT COUNT(*) FROM customers WHERE is_active=1").fetchone()[0],
        'month_treatments': conn.execute("SELECT COUNT(*) FROM treatments WHERE visit_time>=?", (month_start,)).fetchone()[0],
        'pending_followups': conn.execute("SELECT COUNT(*) FROM followups WHERE status='pending'").fetchone()[0],
        'month_income': conn.execute("SELECT COALESCE(SUM(paid_fee),0) FROM treatments WHERE visit_time>=?", (month_start,)).fetchone()[0],
    }
    # 新客户数
    stats['month_new_customers'] = conn.execute(
        "SELECT COUNT(*) FROM customers WHERE created_at>=? AND is_active=1", (month_start,)
    ).fetchone()[0]
    conn.close()
    settings = get_clinic_settings()
    return render_template('index.html',
                           today_followups=today_followups,
                           upcoming_next_visits=upcoming_next_visits,
                           stats=stats,
                           settings=settings,
                           user=get_current_user())


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['real_name'] = user['real_name']
            session['role'] = user['role']
            log_action('登录')
            return redirect(url_for('index'))
        flash('用户名或密码错误', 'error')
    settings = get_clinic_settings()
    return render_template('login.html', settings=settings)


@app.route('/logout')
def logout():
    log_action('退出登录')
    session.clear()
    return redirect(url_for('login'))


# ============================================================
# 客户档案
# ============================================================
@app.route('/customers')
@login_required
def customers():
    settings = get_clinic_settings()
    return render_template('customers.html', settings=settings, user=get_current_user())


@app.route('/customers/<int:cid>')
@login_required
def customer_detail(cid):
    import traceback as _tb
    try:
        conn = get_db()
        customer = conn.execute(
            "SELECT c.*, u.real_name as doctor_name FROM customers c LEFT JOIN users u ON c.first_doctor_id=u.id WHERE c.id=?", (cid,)
        ).fetchone()
        if not customer:
            conn.close()
            flash('客户不存在', 'error')
            return redirect(url_for('customers'))
        treatments = conn.execute(
            "SELECT t.*, u.real_name as doctor_name FROM treatments t LEFT JOIN users u ON t.doctor_id=u.id WHERE t.customer_id=? ORDER BY t.visit_time DESC", (cid,)
        ).fetchall()
        followups = conn.execute(
            "SELECT * FROM followups WHERE customer_id=? ORDER BY plan_date DESC", (cid,)
        ).fetchall()
        conn.close()
        settings = get_clinic_settings()
        return render_template('customer_detail.html',
                               customer=customer,
                               treatments=treatments,
                               followups=followups,
                               settings=settings,
                               user=get_current_user())
    except Exception as _e:
        _err = _tb.format_exc()
        app.logger.error("customer_detail error: %s", _err)
        return f"<pre style='color:red;padding:20px'><b>调试错误信息（请截图发给开发者）：</b>\n\n{_err}</pre>", 500


# ============================================================
# 诊疗记录页
# ============================================================
@app.route('/treatments')
@login_required
def treatments():
    settings = get_clinic_settings()
    return render_template('treatments.html', settings=settings, user=get_current_user())


@app.route('/treatments/<int:tid>')
@login_required
def treatment_detail(tid):
    conn = get_db()
    treatment = conn.execute(
        """SELECT t.*, u.real_name as doctor_name, c.name as customer_name, c.phone as customer_phone
           FROM treatments t
           LEFT JOIN users u ON t.doctor_id=u.id
           LEFT JOIN customers c ON t.customer_id=c.id
           WHERE t.id=?""", (tid,)
    ).fetchone()
    if not treatment:
        conn.close()
        flash('记录不存在', 'error')
        return redirect(url_for('treatments'))
    files = conn.execute("SELECT * FROM treatment_files WHERE treatment_id=?", (tid,)).fetchall()
    conn.close()
    settings = get_clinic_settings()
    return render_template('treatment_detail.html',
                           treatment=treatment,
                           files=files,
                           settings=settings,
                           user=get_current_user())


# ============================================================
# 回访跟进页
# ============================================================
@app.route('/followups')
@login_required
def followups():
    settings = get_clinic_settings()
    return render_template('followups.html', settings=settings, user=get_current_user())


# ============================================================
# 数据统计页
# ============================================================
@app.route('/statistics')
@login_required
def statistics():
    settings = get_clinic_settings()
    return render_template('statistics.html', settings=settings, user=get_current_user())


# ============================================================
# 系统管理页
# ============================================================
@app.route('/system')
@login_required
def system():
    if session.get('role') != 'admin':
        flash('权限不足', 'error')
        return redirect(url_for('index'))
    conn = get_db()
    users = conn.execute("SELECT * FROM users ORDER BY id").fetchall()
    conn.close()
    settings = get_clinic_settings()
    return render_template('system.html', users=users, settings=settings, user=get_current_user())


# ============================================================
# API - 客户档案
# ============================================================
@app.route('/api/customers', methods=['GET'])
@login_required
def api_customers_list():
    q = request.args.get('q', '').strip()
    tag = request.args.get('tag', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page
    conn = get_db()
    where = ["c.is_active=1"]
    params = []
    if q:
        where.append("(c.name LIKE ? OR c.phone LIKE ?)")
        params += [f'%{q}%', f'%{q}%']
    if tag:
        where.append("c.tags LIKE ?")
        params.append(f'%{tag}%')
    sql_where = " AND ".join(where)
    total = conn.execute(f"SELECT COUNT(*) FROM customers c WHERE {sql_where}", params).fetchone()[0]
    rows = conn.execute(
        f"""SELECT c.*, u.real_name as doctor_name,
            (SELECT COUNT(*) FROM treatments t WHERE t.customer_id=c.id) as treatment_count
            FROM customers c LEFT JOIN users u ON c.first_doctor_id=u.id
            WHERE {sql_where} ORDER BY c.created_at DESC LIMIT ? OFFSET ?""",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    data = []
    for r in rows:
        d = dict(r)
        try:
            d['tags'] = json.loads(d['tags'] or '[]')
        except Exception:
            d['tags'] = []
        data.append(d)
    return jsonify({'success': True, 'data': data, 'total': total, 'page': page, 'per_page': per_page})


@app.route('/api/customers', methods=['POST'])
@login_required
def api_customers_create():
    data = request.json or {}
    required = ['name', 'phone']
    for f in required:
        if not data.get(f, '').strip():
            return jsonify({'success': False, 'msg': f'字段 {f} 不能为空'})
    conn = get_db()
    existing = conn.execute("SELECT id FROM customers WHERE phone=?", (data['phone'].strip(),)).fetchone()
    if existing:
        conn.close()
        return jsonify({'success': False, 'msg': '该手机号已存在，请勿重复建档'})
    tags = json.dumps(data.get('tags', []), ensure_ascii=False)
    try:
        c = conn.execute(
            """INSERT INTO customers
               (name, gender, birth_year, phone, wechat, address,
                dental_history, allergy_history, systemic_disease, dental_condition,
                tags, first_doctor_id)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data['name'].strip(), data.get('gender', ''), data.get('birth_year'),
             data['phone'].strip(), data.get('wechat', ''), data.get('address', ''),
             data.get('dental_history', ''), data.get('allergy_history', ''),
             data.get('systemic_disease', ''), data.get('dental_condition', ''),
             tags, session['user_id'])
        )
        conn.commit()
        cid = c.lastrowid
        log_action('新增客户', 'customer', cid, data['name'])
        conn.close()
        return jsonify({'success': True, 'id': cid, 'msg': '建档成功'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'msg': str(e)})


@app.route('/api/customers/<int:cid>', methods=['GET'])
@login_required
def api_customer_get(cid):
    conn = get_db()
    r = conn.execute(
        "SELECT c.*, u.real_name as doctor_name FROM customers c LEFT JOIN users u ON c.first_doctor_id=u.id WHERE c.id=?",
        (cid,)
    ).fetchone()
    conn.close()
    if not r:
        return jsonify({'success': False, 'msg': '客户不存在'})
    d = dict(r)
    try:
        d['tags'] = json.loads(d['tags'] or '[]')
    except Exception:
        d['tags'] = []
    return jsonify({'success': True, 'data': d})


@app.route('/api/customers/<int:cid>', methods=['PUT'])
@login_required
def api_customer_update(cid):
    data = request.json or {}
    conn = get_db()
    existing = conn.execute("SELECT id FROM customers WHERE phone=? AND id!=?",
                            (data.get('phone', '').strip(), cid)).fetchone()
    if existing:
        conn.close()
        return jsonify({'success': False, 'msg': '该手机号已被其他客户使用'})
    tags = json.dumps(data.get('tags', []), ensure_ascii=False)
    conn.execute(
        """UPDATE customers SET
           name=?, gender=?, birth_year=?, phone=?, wechat=?, address=?,
           dental_history=?, allergy_history=?, systemic_disease=?, dental_condition=?,
           tags=?, updated_at=datetime('now','localtime')
           WHERE id=?""",
        (data.get('name', ''), data.get('gender', ''), data.get('birth_year'),
         data.get('phone', '').strip(), data.get('wechat', ''), data.get('address', ''),
         data.get('dental_history', ''), data.get('allergy_history', ''),
         data.get('systemic_disease', ''), data.get('dental_condition', ''),
         tags, cid)
    )
    conn.commit()
    conn.close()
    log_action('编辑客户', 'customer', cid)
    return jsonify({'success': True, 'msg': '更新成功'})


@app.route('/api/customers/<int:cid>', methods=['DELETE'])
@admin_required
def api_customer_delete(cid):
    conn = get_db()
    conn.execute("UPDATE customers SET is_active=0 WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    log_action('删除客户', 'customer', cid)
    return jsonify({'success': True, 'msg': '已禁用'})


@app.route('/api/customers/export')
@login_required
def api_customers_export():
    conn = get_db()
    rows = conn.execute(
        """SELECT c.name, c.gender, c.birth_year, c.phone, c.wechat, c.address,
                  c.dental_history, c.allergy_history, c.systemic_disease, c.dental_condition,
                  c.tags, u.real_name as doctor_name, c.created_at
           FROM customers c LEFT JOIN users u ON c.first_doctor_id=u.id
           WHERE c.is_active=1 ORDER BY c.created_at DESC"""
    ).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '客户档案'
    headers = ['姓名', '性别', '出生年份', '手机号', '微信号', '地址',
               '口腔病史', '药物过敏史', '全身疾病史', '牙齿基础情况', '标签', '首诊医生', '建档时间']
    header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row_data in rows:
        row = list(row_data)
        try:
            row[10] = ', '.join(json.loads(row[10] or '[]'))
        except Exception:
            row[10] = ''
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment;filename=customers.xlsx'})


# ============================================================
# API - 诊疗记录
# ============================================================
@app.route('/api/treatments', methods=['GET'])
@login_required
def api_treatments_list():
    customer_id = request.args.get('customer_id')
    q = request.args.get('q', '').strip()
    doctor_id = request.args.get('doctor_id')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page
    conn = get_db()
    where = ["1=1"]
    params = []
    if customer_id:
        where.append("t.customer_id=?")
        params.append(customer_id)
    if q:
        where.append("(c.name LIKE ? OR c.phone LIKE ? OR t.diagnosis LIKE ?)")
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if doctor_id and session.get('role') == 'doctor':
        where.append("t.doctor_id=?")
        params.append(session['user_id'])
    elif doctor_id:
        where.append("t.doctor_id=?")
        params.append(doctor_id)
    if date_from:
        where.append("t.visit_time>=?")
        params.append(date_from)
    if date_to:
        where.append("t.visit_time<=?")
        params.append(date_to + ' 23:59:59')
    sql_where = " AND ".join(where)
    total = conn.execute(
        f"SELECT COUNT(*) FROM treatments t LEFT JOIN customers c ON t.customer_id=c.id WHERE {sql_where}", params
    ).fetchone()[0]
    rows = conn.execute(
        f"""SELECT t.*, u.real_name as doctor_name, c.name as customer_name, c.phone as customer_phone
            FROM treatments t
            LEFT JOIN users u ON t.doctor_id=u.id
            LEFT JOIN customers c ON t.customer_id=c.id
            WHERE {sql_where} ORDER BY t.visit_time DESC LIMIT ? OFFSET ?""",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    data = []
    for r in rows:
        d = dict(r)
        try:
            d['treatment_items'] = json.loads(d['treatment_items'] or '[]')
        except Exception:
            d['treatment_items'] = []
        data.append(d)
    return jsonify({'success': True, 'data': data, 'total': total})


@app.route('/api/treatments', methods=['POST'])
@login_required
def api_treatments_create():
    data = request.json or {}
    if not data.get('customer_id'):
        return jsonify({'success': False, 'msg': '请选择患者'})
    if not data.get('visit_time'):
        return jsonify({'success': False, 'msg': '请填写就诊时间'})
    conn = get_db()
    items = json.dumps(data.get('treatment_items', []), ensure_ascii=False)
    total_fee = float(data.get('total_fee', 0) or 0)
    paid_fee = float(data.get('paid_fee', 0) or 0)
    unpaid_fee = round(total_fee - paid_fee, 2)
    try:
        c = conn.execute(
            """INSERT INTO treatments
               (customer_id, visit_time, doctor_id, visit_type,
                chief_complaint, exam_result, xray_result, diagnosis,
                treatment_items, treatment_area, treatment_process,
                drug_name, drug_dosage, drug_usage, drug_period,
                total_fee, paid_fee, unpaid_fee,
                advice, contraindications, next_visit_date, next_exam_items, status)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data['customer_id'], data['visit_time'],
             data.get('doctor_id', session['user_id']), data.get('visit_type', '初诊'),
             data.get('chief_complaint', ''), data.get('exam_result', ''),
             data.get('xray_result', ''), data.get('diagnosis', ''),
             items, data.get('treatment_area', ''), data.get('treatment_process', ''),
             data.get('drug_name', ''), data.get('drug_dosage', ''),
             data.get('drug_usage', ''), data.get('drug_period', ''),
             total_fee, paid_fee, unpaid_fee,
             data.get('advice', ''), data.get('contraindications', ''),
             data.get('next_visit_date', ''), data.get('next_exam_items', ''),
             data.get('status', 'done'))
        )
        conn.commit()
        tid = c.lastrowid
        log_action('新增诊疗记录', 'treatment', tid)
        conn.close()
        return jsonify({'success': True, 'id': tid, 'msg': '记录成功'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'msg': str(e)})


@app.route('/api/treatments/<int:tid>', methods=['GET'])
@login_required
def api_treatment_get(tid):
    conn = get_db()
    r = conn.execute(
        """SELECT t.*, u.real_name as doctor_name, c.name as customer_name
           FROM treatments t
           LEFT JOIN users u ON t.doctor_id=u.id
           LEFT JOIN customers c ON t.customer_id=c.id
           WHERE t.id=?""", (tid,)
    ).fetchone()
    files = conn.execute("SELECT * FROM treatment_files WHERE treatment_id=?", (tid,)).fetchall()
    conn.close()
    if not r:
        return jsonify({'success': False, 'msg': '记录不存在'})
    d = dict(r)
    try:
        d['treatment_items'] = json.loads(d['treatment_items'] or '[]')
    except Exception:
        d['treatment_items'] = []
    d['files'] = [dict(f) for f in files]
    return jsonify({'success': True, 'data': d})


@app.route('/api/treatments/<int:tid>', methods=['PUT'])
@login_required
def api_treatment_update(tid):
    conn = get_db()
    t = conn.execute("SELECT * FROM treatments WHERE id=?", (tid,)).fetchone()
    if not t:
        conn.close()
        return jsonify({'success': False, 'msg': '记录不存在'})
    # 医生只能编辑自己的草稿
    if session.get('role') == 'doctor':
        if t['doctor_id'] != session['user_id']:
            conn.close()
            return jsonify({'success': False, 'msg': '只能编辑本人创建的记录'})
    data = request.json or {}
    items = json.dumps(data.get('treatment_items', []), ensure_ascii=False)
    total_fee = float(data.get('total_fee', 0) or 0)
    paid_fee = float(data.get('paid_fee', 0) or 0)
    unpaid_fee = round(total_fee - paid_fee, 2)
    conn.execute(
        """UPDATE treatments SET
           visit_time=?, visit_type=?, chief_complaint=?, exam_result=?, xray_result=?,
           diagnosis=?, treatment_items=?, treatment_area=?, treatment_process=?,
           drug_name=?, drug_dosage=?, drug_usage=?, drug_period=?,
           total_fee=?, paid_fee=?, unpaid_fee=?,
           advice=?, contraindications=?, next_visit_date=?, next_exam_items=?,
           status=?, updated_at=datetime('now','localtime')
           WHERE id=?""",
        (data.get('visit_time'), data.get('visit_type'), data.get('chief_complaint'),
         data.get('exam_result'), data.get('xray_result'), data.get('diagnosis'),
         items, data.get('treatment_area'), data.get('treatment_process'),
         data.get('drug_name'), data.get('drug_dosage'), data.get('drug_usage'), data.get('drug_period'),
         total_fee, paid_fee, unpaid_fee,
         data.get('advice'), data.get('contraindications'), data.get('next_visit_date'), data.get('next_exam_items'),
         data.get('status', 'done'), tid)
    )
    conn.commit()
    conn.close()
    log_action('编辑诊疗记录', 'treatment', tid)
    return jsonify({'success': True, 'msg': '更新成功'})


@app.route('/api/treatments/<int:tid>/upload', methods=['POST'])
@login_required
def api_treatment_upload(tid):
    if 'file' not in request.files:
        return jsonify({'success': False, 'msg': '未选择文件'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'msg': '未选择文件'})
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'msg': '不支持该文件类型'})
    ext = file.filename.rsplit('.', 1)[1].lower()
    ts = datetime.now().strftime('%Y%m%d%H%M%S%f')
    saved_name = f"t{tid}_{ts}.{ext}"
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_name))
    conn = get_db()
    conn.execute(
        "INSERT INTO treatment_files(treatment_id,filename,original_name,file_type) VALUES(?,?,?,?)",
        (tid, saved_name, file.filename, ext)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'filename': saved_name, 'original_name': file.filename})


@app.route('/api/treatments/export')
@login_required
def api_treatments_export():
    customer_id = request.args.get('customer_id')
    conn = get_db()
    where = "1=1"
    params = []
    if customer_id:
        where = "t.customer_id=?"
        params.append(customer_id)
    rows = conn.execute(
        f"""SELECT c.name as 患者姓名, c.phone as 手机号,
                   t.visit_time as 就诊时间, t.visit_type as 就诊类型,
                   u.real_name as 接诊医生, t.diagnosis as 诊断结论,
                   t.treatment_items as 治疗项目, t.treatment_area as 治疗部位,
                   t.total_fee as 总费用, t.paid_fee as 已付, t.unpaid_fee as 未付,
                   t.next_visit_date as 建议复诊日期, t.advice as 医嘱
            FROM treatments t
            LEFT JOIN customers c ON t.customer_id=c.id
            LEFT JOIN users u ON t.doctor_id=u.id
            WHERE {where} ORDER BY t.visit_time DESC""", params
    ).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '诊疗记录'
    headers = list(rows[0].keys()) if rows else []
    header_fill = PatternFill(start_color='43A047', end_color='43A047', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row_data in rows:
        row = list(row_data)
        try:
            idx = headers.index('治疗项目')
            row[idx] = ', '.join(json.loads(row[idx] or '[]'))
        except Exception:
            pass
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment;filename=treatments.xlsx'})


# ============================================================
# API - 回访跟进
# ============================================================
@app.route('/api/followups', methods=['GET'])
@login_required
def api_followups_list():
    status = request.args.get('status', '')
    customer_id = request.args.get('customer_id', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page
    conn = get_db()
    where = ["1=1"]
    params = []
    if status:
        where.append("f.status=?")
        params.append(status)
    if customer_id:
        where.append("f.customer_id=?")
        params.append(customer_id)
    if date_from:
        where.append("f.plan_date>=?")
        params.append(date_from)
    if date_to:
        where.append("f.plan_date<=?")
        params.append(date_to)
    sql_where = " AND ".join(where)
    total = conn.execute(
        f"SELECT COUNT(*) FROM followups f WHERE {sql_where}", params
    ).fetchone()[0]
    rows = conn.execute(
        f"""SELECT f.*, c.name as customer_name, c.phone as customer_phone
            FROM followups f
            LEFT JOIN customers c ON f.customer_id=c.id
            WHERE {sql_where} ORDER BY f.plan_date ASC, f.id DESC LIMIT ? OFFSET ?""",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(r) for r in rows], 'total': total})


@app.route('/api/followups', methods=['POST'])
@login_required
def api_followups_create():
    data = request.json or {}
    if not data.get('customer_id'):
        return jsonify({'success': False, 'msg': '请选择患者'})
    if not data.get('plan_date'):
        return jsonify({'success': False, 'msg': '请选择回访日期'})
    conn = get_db()
    try:
        c = conn.execute(
            """INSERT INTO followups(customer_id, treatment_id, plan_date, followup_type, plan_content, created_by)
               VALUES(?,?,?,?,?,?)""",
            (data['customer_id'], data.get('treatment_id'), data['plan_date'],
             data.get('followup_type', '术后回访'), data.get('plan_content', ''),
             session['user_id'])
        )
        conn.commit()
        fid = c.lastrowid
        conn.close()
        log_action('创建回访计划', 'followup', fid)
        return jsonify({'success': True, 'id': fid, 'msg': '回访计划创建成功'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'msg': str(e)})


@app.route('/api/followups/<int:fid>', methods=['PUT'])
@login_required
def api_followup_update(fid):
    data = request.json or {}
    conn = get_db()
    conn.execute(
        """UPDATE followups SET
           status=?, actual_date=?, result=?, recovery=?,
           next_appointment=?, next_followup_needs=?
           WHERE id=?""",
        (data.get('status', 'done'), data.get('actual_date', date.today().isoformat()),
         data.get('result', ''), data.get('recovery', ''),
         1 if data.get('next_appointment') else 0,
         data.get('next_followup_needs', ''), fid)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '回访记录已保存'})


# ============================================================
# API - 数据统计
# ============================================================
@app.route('/api/statistics/overview')
@login_required
def api_stats_overview():
    conn = get_db()
    today = date.today().isoformat()
    month_start = today[:7] + '-01'
    year_start = today[:4] + '-01-01'
    result = {
        'total_customers': conn.execute("SELECT COUNT(*) FROM customers WHERE is_active=1").fetchone()[0],
        'total_treatments': conn.execute("SELECT COUNT(*) FROM treatments").fetchone()[0],
        'month_new_customers': conn.execute("SELECT COUNT(*) FROM customers WHERE created_at>=? AND is_active=1", (month_start,)).fetchone()[0],
        'month_treatments': conn.execute("SELECT COUNT(*) FROM treatments WHERE visit_time>=?", (month_start,)).fetchone()[0],
        'month_income': conn.execute("SELECT COALESCE(SUM(paid_fee),0) FROM treatments WHERE visit_time>=?", (month_start,)).fetchone()[0],
        'pending_followups': conn.execute("SELECT COUNT(*) FROM followups WHERE status='pending'").fetchone()[0],
        'today_followups': conn.execute("SELECT COUNT(*) FROM followups WHERE plan_date=? AND status='pending'", (today,)).fetchone()[0],
    }
    # 近12个月接诊趋势
    monthly_data = conn.execute(
        """SELECT strftime('%Y-%m', visit_time) as month, COUNT(*) as cnt, COALESCE(SUM(paid_fee),0) as income
           FROM treatments WHERE visit_time>=? GROUP BY month ORDER BY month""", (year_start,)
    ).fetchall()
    result['monthly_trend'] = [dict(r) for r in monthly_data]
    # 治疗项目统计（本年）
    all_treatments = conn.execute(
        "SELECT treatment_items FROM treatments WHERE visit_time>=?", (year_start,)
    ).fetchall()
    item_counter = {}
    for row in all_treatments:
        try:
            items = json.loads(row[0] or '[]')
            for item in items:
                item_counter[item] = item_counter.get(item, 0) + 1
        except Exception:
            pass
    sorted_items = sorted(item_counter.items(), key=lambda x: -x[1])[:10]
    result['top_items'] = [{'name': k, 'count': v} for k, v in sorted_items]
    # 新老客户比
    visit_types = conn.execute(
        """SELECT visit_type, COUNT(*) as cnt FROM treatments WHERE visit_time>=? GROUP BY visit_type""", (month_start,)
    ).fetchall()
    result['visit_types'] = [dict(r) for r in visit_types]
    conn.close()
    return jsonify({'success': True, 'data': result})


# ============================================================
# API - 系统管理
# ============================================================
@app.route('/api/users', methods=['GET'])
@login_required
def api_users_list():
    conn = get_db()
    rows = conn.execute("SELECT id, username, real_name, role, is_active, created_at FROM users ORDER BY id").fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})


@app.route('/api/users', methods=['POST'])
@admin_required
def api_users_create():
    data = request.json or {}
    if not data.get('username') or not data.get('password') or not data.get('real_name'):
        return jsonify({'success': False, 'msg': '用户名、密码、姓名不能为空'})
    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE username=?", (data['username'],)).fetchone()
    if existing:
        conn.close()
        return jsonify({'success': False, 'msg': '用户名已存在'})
    conn.execute("INSERT INTO users(username,password_hash,real_name,role) VALUES(?,?,?,?)",
                 (data['username'], generate_password_hash(data['password']),
                  data['real_name'], data.get('role', 'doctor')))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '账号创建成功'})


@app.route('/api/users/<int:uid>', methods=['PUT'])
@admin_required
def api_user_update(uid):
    data = request.json or {}
    conn = get_db()
    if data.get('password'):
        conn.execute("UPDATE users SET real_name=?,role=?,is_active=?,password_hash=? WHERE id=?",
                     (data.get('real_name', ''), data.get('role', 'doctor'),
                      1 if data.get('is_active', True) else 0,
                      generate_password_hash(data['password']), uid))
    else:
        conn.execute("UPDATE users SET real_name=?,role=?,is_active=? WHERE id=?",
                     (data.get('real_name', ''), data.get('role', 'doctor'),
                      1 if data.get('is_active', True) else 0, uid))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '更新成功'})


@app.route('/api/settings', methods=['GET'])
@login_required
def api_settings_get():
    return jsonify({'success': True, 'data': get_clinic_settings()})


@app.route('/api/settings', methods=['POST'])
@admin_required
def api_settings_update():
    data = request.json or {}
    conn = get_db()
    for key, val in data.items():
        conn.execute("INSERT OR REPLACE INTO clinic_settings(key,value) VALUES(?,?)", (key, val))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '保存成功'})


@app.route('/api/dict/<category>', methods=['GET'])
@login_required
def api_dict_get(category):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM dict_items WHERE category=? AND is_active=1 ORDER BY sort_order, id", (category,)
    ).fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})


@app.route('/api/dict', methods=['POST'])
@admin_required
def api_dict_create():
    data = request.json or {}
    conn = get_db()
    conn.execute("INSERT INTO dict_items(category,name,sort_order) VALUES(?,?,?)",
                 (data['category'], data['name'], data.get('sort_order', 0)))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '添加成功'})


@app.route('/api/dict/<int:did>', methods=['DELETE'])
@admin_required
def api_dict_delete(did):
    conn = get_db()
    conn.execute("UPDATE dict_items SET is_active=0 WHERE id=?", (did,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '已删除'})


@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.json or {}
    old_pwd = data.get('old_password', '')
    new_pwd = data.get('new_password', '')
    if not old_pwd or not new_pwd:
        return jsonify({'success': False, 'msg': '请填写完整'})
    if len(new_pwd) < 6:
        return jsonify({'success': False, 'msg': '新密码至少6位'})
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (session['user_id'],)).fetchone()
    if not check_password_hash(user['password_hash'], old_pwd):
        conn.close()
        return jsonify({'success': False, 'msg': '原密码错误'})
    conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                 (generate_password_hash(new_pwd), session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'msg': '密码修改成功'})


@app.route('/api/backup')
@admin_required
def api_backup():
    """备份数据库"""
    if not os.path.exists(DB_PATH):
        return jsonify({'success': False, 'msg': '数据库文件不存在'})
    with open(DB_PATH, 'rb') as f:
        data = f.read()
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return Response(data,
                    mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment;filename=dental_backup_{ts}.db'})


@app.route('/api/doctors')
@login_required
def api_doctors():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, real_name, username FROM users WHERE is_active=1 ORDER BY role DESC, id"
    ).fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})


# 静态文件服务（上传文件）
@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


if __name__ == '__main__':
    init_db()
    print("=" * 50)
    print("口腔诊所轻量化管理系统启动中...")
    print("访问地址: http://localhost:5000")
    print("默认账号: admin  密码: admin123")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)
